from tkinter import ttk, messagebox
from tkinter import filedialog as fd

from tkinter import *

from os import startfile, listdir, path, mkdir

from function.main import *

from model.color import Colors
from model.fonts import Fonts

from assets.icon import icon as mylan_icon

from PIL import ImageTk, Image
import clipboard
import csv
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

PADX = 10
IPX = 4
PADY = 5
IPY = 2

class Window(Tk):
	def __init__(self):
		super().__init__()
		self.title("RFID - Value converter")

		photo = PhotoImage( data=mylan_icon )
		self.wm_iconphoto(False, photo)


		self.colors = Colors()
		self.fonts  = Fonts()

		self.ascii_1 = StringVar()
		self.ascii_1_cat = StringVar()
		self.ascii_1.trace('w', self.ascii_1_on_change)

		self.ascii_2 = StringVar()
		self.hexa_1  = StringVar()
		self.hexa_2 = StringVar()
		self.hexa_2.trace('w', self.hexa_2_on_change)



		self.tabbed_view = ttk.Notebook(self)

		# self.stylingTableMenu()

		self.tab_1 = Frame( self.tabbed_view )
		self.tab_2 = Frame( self.tabbed_view )
		self.tab_3 = Frame( self.tabbed_view )
		self.tab_4 = Frame( self.tabbed_view )

		self.tabbed_view.add( self.tab_1, text="    Ascii to Hex    " )
		self.tabbed_view.add( self.tab_2, text="    Hex to Ascii    " )
		self.tabbed_view.add( self.tab_3, text="    Excel parser    " )
		self.tabbed_view.add( self.tab_4, text="    Duplicate checker" )

		self.headFrame = LabelFrame(self.tab_3, text="File".upper())

		self.file_path = StringVar()
		self.file_name = StringVar()
		self.currentSheet = StringVar()
		self.file_name.set("No file has been selected")
		self.wb = None

		self.currentSheet.trace("w", self.comboBoxChangeSheet )

		self.columns = []
		self.df = []
		self.treeLabel = StringVar()
		self.sheetSelection = ttk.Combobox( self.headFrame, textvariable = self.currentSheet, width=10 );

		self.source_var = StringVar()
		self.result_var = StringVar()

		self.tab_3_footer_frame = LabelFrame( self.tab_3, text="CONFIG" )

		self.source_column = ttk.Combobox( self.tab_3_footer_frame, textvariable=self.source_var )
		self.result_column = ttk.Combobox( self.tab_3_footer_frame, textvariable=self.result_var )

		self.source_column.bind("<FocusIn>", lambda e: e.widget.master.focus_set())
		self.result_column.bind("<FocusIn>", lambda e: e.widget.master.focus_set())


		self.check_folder_path = StringVar()
		self.check_header = StringVar()
		self.file_in_progress = StringVar()
		self.checking_progress = StringVar()

		self.init_tab_1()
		self.init_tab_2()
		self.init_tab_3()
		self.init_tab_4()

		self.tabbed_view.pack( expand = 1, fill=BOTH )


	def stylingTableMenu(self):
		style = ttk.Style()

		style.theme_create(
			"tabbed_view", parent="alt",
			settings = {
		        "TNotebook": {
		        	"configure": {
		        		"tabmargins": [2, 5, 2, 0],
		        	},
		        },
		        "TNotebook.Tab": {
		            "configure": {
		            	"padding": [15, 1],
		            	"font": self.fonts.normal,
		            	"background": self.colors.gray,
		            	"borderwidth": 0,
		            	"focuscolor": self.colors.white,
		            },
		            "map":{
		            	"background": [ ("selected", "#ffffff")],
		            }
		        },
		        "Treeview": {
		         	"configure": {
		         		"background": self.colors.white,
		         		"fieldbackground": self.colors.white,
						"font": self.fonts.small
		         	}
		        },
		        "Treeview.Heading": {
		         	"configure":{
		         		"background": self.colors.cyan,
		         		"foreground": self.colors.white,
						"font": self.fonts.small_bold
		         	}
		        },
		        "TCombobox.Listbox": {
		        	"configure": {

		        	}
		        }
		    })
		style.configure("Tab", focuscolor=style.configure(".")["background"])
		style.theme_use("tabbed_view")


	def init_tab_1(self):
		# self.label_1 = Label( self.tab_1, text="Unikey có thể sẽ làm phần mềm bị lỗi, vui lòng tắt nó đi trước khi dùng nhé" )
		# self.label_1.pack( side=TOP, fill=BOTH, padx=PADX, pady=PADY )
		self.tab_1_ui_init()

	def ascii_1_on_change(self, *args):
		ascii_string = self.ascii_1.get()
		ascii_list = char_to_bin(ascii_string)
		dec_list = [ binaryToDecimal(x) for x in ascii_list ]
		hex_list = [ decToHexa(x) for x in dec_list ]

		self.delete_detail_tree_view_1_children()
		for i in range( len(ascii_list) ):
			self.detail_tree_view_1.insert(
					parent="",
					index=i,
					iid=i,
					values=( ascii_string[i], ascii_list[i], hex_list[i])
				)

		self.hexa_1.set( ' '.join(hex_list) )
		self.detail_tree_view_1.yview_moveto(1)


	def clipboardHex(self):
		clipboard.copy(self.hexa_1.get())

	def clipboardAscii(self):
		clipboard.copy(self.ascii_2.get())


	def delete_detail_tree_view_1_children(self):
		for item in self.detail_tree_view_1.get_children():
			self.detail_tree_view_1.delete(item)

	def ascii_key_down(self, event):
		return


	def tab_1_ui_init(self):
		frame_0_1 = LabelFrame( self.tab_1, text="ASCII" )
		# Label( frame_0_1, text="ASCII", font=self.fonts.normal).pack( side=LEFT, padx=PADX, pady=PADY )
		frame_0_1.pack( side=TOP, fill=X, padx=PADX, pady=PADY )

		self.tab_1_ascii_entry = Entry(
				frame_0_1,
				textvariable = self.ascii_1,
				validate = "key",
				font=self.fonts.normal,
			)
		# self.tab_1_ascii_entry.bind("<Key>", self.ascii_key_down)
		self.tab_1_ascii_entry.pack(side=TOP, fill=BOTH, padx=PADX, pady=PADY )

		frame_0_2 = LabelFrame( self.tab_1, text="HEXA" )
		# Label( frame_0_2, text="HEXA", font=self.fonts.normal).pack( side=LEFT, padx=PADX, pady=PADY )
		frame_0_2.pack( side=TOP, fill=X, padx=PADX, pady=PADY )

		frame_0_3 = Frame( frame_0_2 )
		self.tab_1_hex_entry = Entry(
				frame_0_3,
				textvariable = self.hexa_1,
				validate = "key",
				font=self.fonts.normal,
				state=DISABLED
			)
		self.tab_1_hex_entry.pack(side=LEFT, fill=BOTH, expand=True, padx=PADX, pady=PADY )
		Button( frame_0_3, text="Copy", command=self.clipboardHex,
			borderwidth = 0,
			background = self.colors.white,
			font = self.fonts.x_small,
			activebackground = self.colors.white
		).pack( side=LEFT, padx=PADX, pady=PADY, ipadx=15, ipady=4 )
		frame_0_3.pack( side=TOP, fill=X )


		tree_frame = LabelFrame( self.tab_1, text="DETAIL" )

		self.detail_tree_view_1 = ttk.Treeview( tree_frame, column=(0, 1, 2), height = 12, show="headings",)
		self.detail_tree_view_1.heading(0, text="Ascii", anchor="w")
		self.detail_tree_view_1.column(0, anchor="w")
		self.detail_tree_view_1.heading(1, text="Binary", anchor="w")
		self.detail_tree_view_1.column(1, anchor="w")
		self.detail_tree_view_1.heading(2, text="Hexa", anchor="w")
		self.detail_tree_view_1.column(2, anchor="w")

		verscrlbar = ttk.Scrollbar(tree_frame,
                           orient ="vertical",
                           command = self.detail_tree_view_1.yview)
		verscrlbar.pack( side=RIGHT, fill=Y )
		self.detail_tree_view_1.configure(yscrollcommand=verscrlbar.set)
		self.detail_tree_view_1.pack( side=TOP, fill=BOTH, padx=PADX, pady=PADY )
		tree_frame.pack( side=TOP, fill=BOTH, padx=PADX, pady=PADY )

	def init_tab_2(self):

		# self.label_2 = Label( self.tab_2, text="Những ký tự bên ngoài [1-9][A-Z] và <space/> đều được tính là không hợp lệ" )
		# self.label_2.pack( side=TOP, fill=BOTH, padx=PADX, pady=PADY )

		self.tab_2_ui_init()

	def tab_2_ui_init(self):
		frame_0_1 = LabelFrame( self.tab_2, text="HEXA" )
		# Label( frame_0_1, text="HEXA", font=self.fonts.normal).pack( side=LEFT, padx=PADX, pady=PADY )
		frame_0_1.pack( side=TOP, fill=X, padx=PADX, pady=PADY )

		self.tab_2_ascii_entry = Entry(
				frame_0_1,
				textvariable = self.hexa_2,
				validate = "key",
				font=self.fonts.normal,
			)
		# self.tab_2_ascii_entry.bind("<Key>", self.ascii_key_down)
		self.tab_2_ascii_entry.pack(side=TOP, fill=BOTH, padx=PADX, pady=PADY )

		frame_0_2 = LabelFrame( self.tab_2, text="ASCII" )
		# Label( frame_0_2, text="ASCII", font=self.fonts.normal).pack( side=LEFT, padx=PADX, pady=PADY )
		frame_0_2.pack( side=TOP, fill=X, padx=PADX, pady=PADY )

		frame_0_3 = Frame( frame_0_2 )
		self.tab_2_hex_entry = Entry(
				frame_0_3,
				textvariable = self.ascii_2,
				validate = "key",
				font=self.fonts.normal,
				state=DISABLED
			)
		self.tab_2_hex_entry.pack(side=LEFT, fill=BOTH, expand=True, padx=PADX, pady=PADY )
		Button( frame_0_3, text="Copy", command=self.clipboardAscii,
			borderwidth = 0,
			background = self.colors.white,
			font = self.fonts.x_small,
			activebackground = self.colors.white
		).pack( side=LEFT, padx=PADX, pady=PADY, ipadx=15, ipady=4 )
		frame_0_3.pack( side=TOP, fill=X )


		tree_frame = LabelFrame( self.tab_2, text="DETAIL" )

		self.detail_tree_view_2 = ttk.Treeview( tree_frame, column=(0, 1, 2), height = 12, show="headings",)
		self.detail_tree_view_2.heading(0, text="Hexa", anchor="w")
		self.detail_tree_view_2.column(0, anchor="w")
		self.detail_tree_view_2.heading(1, text="Ascii", anchor="w")
		self.detail_tree_view_2.column(1, anchor="w")
		self.detail_tree_view_2.heading(2, text="Binary", anchor="w")
		self.detail_tree_view_2.column(2, anchor="w")


		verscrlbar = ttk.Scrollbar(tree_frame,
                           orient ="vertical",
                           command = self.detail_tree_view_2.yview)
		verscrlbar.pack( side=RIGHT, fill=Y )
		self.detail_tree_view_2.configure(yscrollcommand=verscrlbar.set)
		self.detail_tree_view_2.pack( side=TOP, fill=BOTH, padx=PADX, pady=PADY )
		tree_frame.pack( side=TOP, fill=BOTH, padx=PADX, pady=PADY )


	def hex_validate(self, string):
		hex_chars = [ x for x in "0123456789ABCDEF " ]
		valid = True
		for i in string:
			if not i in hex_chars:
				valid = False
		return valid

	def delete_detail_tree_view_2_children(self):
		for item in self.detail_tree_view_2.get_children():
			self.detail_tree_view_2.delete(item)


	def hexa_2_on_change(self, *args):
		hexa_string = self.hexa_2.get()
		valid = self.hex_validate( hexa_string )
		if valid:
			trim_str = hexa_string.replace(" ", "")
			hex_pair = []
			if len(trim_str ) % 2 == 0:
				for i in range(0, len(trim_str ), 2 ):
					hex_pair.append( trim_str[i]+trim_str[i+1] )
			else:
				for i in range( len(trim_str) - 1, 0, -2 ):
					hex_pair.insert(0, trim_str[i-1]+trim_str[i] )
				hex_pair.insert(0, "0"+trim_str[0])

			self.delete_detail_tree_view_2_children()

			int_list = [ int( x, 16 ) for x in hex_pair ]
			char_list = [ chr( x ) for x in int_list ]
			joined_char_list = "".join( char_list )
			bin_list = char_to_bin( joined_char_list )

			for i in range( len( int_list ) ):
				self.detail_tree_view_2.insert( parent ="", index = i, iid=i,
					values = ( hex_pair[i], char_list[i], bin_list[i] )
				)

			self.ascii_2.set( ''.join(char_list) )
			self.detail_tree_view_2.yview_moveto(1)


		else:
			messagebox.showerror(title = "Error", message="Some characters conflict based-16 formal value")


	def initHeader(self):

		self.file_path_entry = Label(
							self.headFrame,
							textvar=self.file_name,
							font=self.fonts.normal
						).pack( side=LEFT, fill=X, padx=PADX, pady=PADY, ipadx=IPX, ipady=IPY,expand=1 )
		self.file_button = Button(
							self.headFrame,
							text = "Browse",
							bg = self.colors.white,
							borderwidth = 0,
							fg= self.colors.black,
							font = self.fonts.small,
							command = self.askExcelFile
						).pack( side=LEFT, padx=PADX, pady=PADY, ipadx=25 )

		self.sheetSelection.config(font=self.fonts.small)
		self.sheetSelection.pack(side =LEFT, padx=PADX, pady=PADY, ipadx=25 )
		self.sheetSelection.bind("<FocusIn>", lambda e: e.widget.master.focus_set())
		self.headFrame.pack( side=TOP, fill=X, padx=2*PADX )

	def initBody( self ):
		self.bodyFrame = LabelFrame( self.tab_3, text="DATA" )
		self.tree = ttk.Treeview(self.bodyFrame, columns=(1, 2, 3), height = 10)
		self.initTree()
		self.bodyFrame.pack( side=TOP, fill=X, padx= 2* PADX, pady=2*PADY )

	def initTree(self):
		self.tree.heading(1, anchor=CENTER, text="#" )
		self.tree.heading(2, anchor=CENTER, text="Somecode")
		self.tree.column(1, anchor=CENTER)
		self.tree.column(2, anchor=CENTER)

		verscrlbar = ttk.Scrollbar(self.bodyFrame,
                           orient ="vertical",
                           command = self.tree.yview)
		verscrlbar.pack( side=RIGHT, fill=Y )
		self.tree.configure(yscrollcommand=verscrlbar.set)

		self.tree.insert( parent="", index=0, iid=0, values=(1, "###") )
		self.tree.pack( side=BOTTOM,fill=X, expand=1, padx=2*PADX, pady=PADY )


	def comboBoxChangeSheet(self, *args):
		self.retriveSheetName(self.currentSheet.get())

	def askExcelFile(self):
		filetype = (
		        ('Excel file', '*.xlsx'),
		        # ('CSV file', '*.csv'),
		        ('All files', '*.*')
   		)
		filename = fd.askopenfilenames(
		        title='Open file',
		        initialdir='/',
		        filetypes=filetype
	        )
		if len( filename ) > 0 :
			file = filename[0]

			extension = file.split('.')[-1]
			if extension == "xlsx":
				self.file_path.set(file)
				self.file_name.set( file.split("/")[-1] )
				self.wb = openpyxl.load_workbook(file)
				self.retriveSheetName( self.wb.sheetnames[0] )
				self.sheetSelection["values"] = self.wb.sheetnames
			else:
				messagebox.showerror(title="Error", message="Selected file has wrong file extension!")

	def clear_treeview(self):
		# this may cause some bug
		for child in self.tree.get_children():
			self.tree.delete(child)

	def re_make_df( self, columns, data ):
		self.columns = columns
		self.df = []
		for row in data:
			tmp = {}
			for i in range(0, len(row)):
				tmp.setdefault(columns[i], row[i])

			self.df.append(tmp)

	def retriveSheetName(self, sheet):

		xl = pd.ExcelFile(self.file_path.get())
		df = xl.parse(sheet)
		self.clear_treeview()
		columns = list(df.columns)
		self.tree["column"] = columns
		self.tree["show"] = "headings"

		for col in self.tree["column"]:
			self.tree.heading(col, text=col, anchor="w")

		df_rows = df.to_numpy().tolist()

		row_index = 1
		indexed_df_rows = []
		for row in df_rows:
			self.tree.insert("", "end", values=row )
			indexed_df_rows.append([ *row ])
			row_index += 1

		self.tree.pack( fill=X, expand=1, side=LEFT )
		self.currentSheet.set(sheet)

		self.source_column["values"] = columns
		self.result_column["values"] = columns

		self.re_make_df(columns, indexed_df_rows)
		# self.treeLabel.set("Hiển thị  kết quả".format( len(df_rows) ))
		self.bodyFrame.configure(text = "Data ( {0} rows found )".format( len(df_rows))	)


	def init_tab_3(self):
		self.initHeader()
		self.initBody()
		self.initFooter()

	def initFooter(self):
		Label( self.tab_3_footer_frame, text="SOURCE COLUMN", font=self.fonts.x_small ).pack( side=LEFT, padx=PADX, pady=PADY )
		self.source_column.pack(side=LEFT, fill=X, padx=PADX, pady=PADY, ipady=3)
		# Label( self.tab_3_footer_frame, text="RESULT COLUMN", font=self.fonts.x_small ).pack( side=LEFT, padx=PADX, pady=PADY )
		# self.result_column.pack(side=LEFT, fill=X, padx=PADX, pady=PADY, ipady=3)
		Button( self.tab_3_footer_frame, text="GENERATE", command=self.generate_code,
			borderwidth = 0,
			background = self.colors.cyan,
			foreground = self.colors.white,
			font = self.fonts.x_small,
			activebackground = self.colors.black,
			activeforeground = self.colors.white
		).pack( side=RIGHT, padx=PADX, pady=PADY, ipadx=15, ipady=0 )
		self.tab_3_footer_frame.pack( side=TOP, fill=X, padx=2*PADX, pady=PADY )

	def generate_code(self):
		source_column = self.source_column.get()
		result_column = self.result_column.get()

		if not source_column:
			messagebox.showerror( title="Error", message = "Source column is required" )
			return

		if source_column == result_column:
			messagebox.showerror( title="Error", message = "Source and result must be defferent" )
			return

		files = [('MS Excel thư quý vị và các bạn', '*.xlsx')]
		try:
			file = fd.asksaveasfile(filetypes = files, defaultextension = files)
		except:
			messagebox.showerror(title="Lỗi", message="Tệp đang mở ở một tác vụ khác, không thể thao tác!")
			return

		file_name = file.name

		wb = openpyxl.Workbook()
		sheet = wb.active

		index_source = self.columns.index( source_column )
		column = self.columns[ index_source ]
		ascii_list = []
		columns = [*self.columns, "hexa_result"]
		header = sheet['A1:{0}1'.format( get_column_letter( len(columns) + 1 ) )]
		
		for index in range(len(columns)):
			header[0][index].value = columns[index]

		body = sheet['A2:{0}{1}'.format( get_column_letter( len(columns) + 1 ), len( self.df ) + 1 )]
		for index in range(len(self.df)):
			frame = self.df[index]
			trimmed_data = frame[column].strip()
			ascii_list = char_to_bin(trimmed_data)
			dec_list = [ binaryToDecimal(x) for x in ascii_list ]
			hex_list = [ decToHexa(x) for x in dec_list ]
			result =" ".join( hex_list )
			frame["hexa_result"] = result

			for col in range(len(columns)):
				key = columns[col]
				body[index][col].value = frame[key]

		wb.save(file_name)

		start_file = messagebox.askyesno( title="Thành công", message="Mở tệp đã lưu?" )
		if start_file:
			startfile(file_name)

	def init_tab_4(self):
		self.paths = []

		select_button_frame = LabelFrame( self.tab_4, text="FOLDER" )
		file_path_entry = Entry( 
							select_button_frame, 							
							font=self.fonts.normal,
							# state="disabled",
							textvariable=self.check_folder_path,
						).pack( side=LEFT, fill=X, padx=PADX, pady=PADY*2, ipadx=IPX, ipady=IPY,expand=1 )
		file_button = Button( 
							select_button_frame, 
							text = "Browse",
							bg = self.colors.white,
							borderwidth = 0,
							fg= self.colors.black,							
							font = self.fonts.small,
							command=self.selectCheckFolder							
						).pack( side=LEFT, padx=PADX, pady=PADY, ipadx=25 )
		activate_button = Button( 
							select_button_frame, 
							text = "Load",
							bg = self.colors.white,
							borderwidth = 0,
							fg= self.colors.black,							
							font = self.fonts.small,
							command=self.load_check_folder						
						).pack( side=LEFT, padx=PADX, pady=PADY, ipadx=25 )
		select_button_frame.pack( side=TOP, fill=X, padx=PADX, pady=PADY )

		self.trees_frame = Frame(self.tab_4, width=100)

		self.left_frame = LabelFrame(self.trees_frame, text="FILES", width=225)
		self.check_tree_left = ttk.Treeview(self.left_frame, columns=(1, 2, 3), height = 10, show="headings" )
		self.check_tree_left.pack(side=LEFT,  fill=BOTH)
		self.left_frame.pack(side=LEFT, fill=BOTH, padx=10)
		self.initLeftTree()

		self.middle_frame = Frame( self.trees_frame, width=50 )
		self.middle_frame.pack(side=LEFT, expand=1, fill=Y)

		move_one_btn = Button(self.middle_frame, 
			text=">",
			bg = self.colors.white,
			borderwidth = 0,
			fg= self.colors.black,							
			font = self.fonts.x_small,
			command=self.move_selected
		)
		move_one_btn.pack(side=TOP, fill=BOTH, padx=5, ipadx=20, pady=15)

		move_all_btn = Button(self.middle_frame, 
			text=">>",
			bg = self.colors.white,
			borderwidth = 0,
			fg= self.colors.black,							
			font = self.fonts.x_small,
			command=self.move_all
		)
		move_all_btn.pack(side=TOP, fill=BOTH, padx=5, ipadx=20, pady=15)

		reverse_one_btn = Button(self.middle_frame, 
			text="<",
			bg = self.colors.white,
			borderwidth = 0,
			fg= self.colors.black,							
			font = self.fonts.x_small,
			command=self.reverse_one
		)
		reverse_one_btn.pack(side=TOP, fill=BOTH, padx=5, ipadx=20, pady=15)

		reverse_all_btn = Button(self.middle_frame, 
			text="<<",
			bg = self.colors.white,
			borderwidth = 0,
			fg= self.colors.black,							
			font = self.fonts.x_small,
			command=self.reverse_all
		)
		reverse_all_btn.pack(side=TOP, fill=BOTH, padx=5, ipadx=20, pady=15)




		self.right_frame = LabelFrame(self.trees_frame, text="CHECK LIST", width=225)
		self.check_tree_right = ttk.Treeview(self.right_frame, columns=(1, 2, 3), height = 10, show="headings" )
		self.check_tree_right.pack(side=LEFT, fill=BOTH)
		self.right_frame.pack(side=RIGHT, fill=BOTH, padx=10)
		self.initRightTree()

		self.trees_frame.pack(side = TOP)

		header_frame = Frame(self.tab_4)
		entry_frame = Frame(header_frame)
		file_path_entry = Entry( 
			entry_frame, 							
			font=self.fonts.normal,
			# state="disabled",
			textvariable=self.check_header,
		).pack( side=LEFT, fill=X, padx=PADX, pady=PADY*2, ipadx=IPX, ipady=IPY,expand=1 )
		entry_frame.pack(side=LEFT, expand=1, fill=X)

		run_button = Button( 
			header_frame, 
			text = "START CHECKING",
			bg = self.colors.cyan,
			borderwidth = 0,
			fg= self.colors.white,							
			font = self.fonts.small,
			command=self.startChecking						
		).pack( side=LEFT, padx=PADX, pady=PADY, ipadx=25 )

		header_frame.pack(side=TOP, fill=X, padx=10)

		checking_progress = Frame(self.tab_4)		
		progress = Label(checking_progress, textvariable=self.checking_progress, fg=self.colors.cyan, font=self.fonts.small)
		checking_file = Label(checking_progress, textvariable=self.file_in_progress, fg=self.colors.cyan, font=self.fonts.small)

		progress.pack(side=LEFT, padx=10, pady=10)
		checking_file.pack(side=LEFT, padx=10, pady=10)
		checking_progress.pack(side=TOP, pady=10, fill=BOTH)

	def selectCheckFolder(self):
		folder = fd.askdirectory()
		self.check_folder_path.set(folder)
		self.load_check_folder()

	def initLeftTree(self):
		self.check_tree_left.heading(1, anchor=CENTER, text="#" )
		self.check_tree_left.heading(2, anchor=W, text="File name")
		self.check_tree_left.heading(3, anchor=CENTER, text="Type")
		self.check_tree_left.column(1, stretch=NO, width=25,anchor=CENTER)
		self.check_tree_left.column(2, anchor=W)
		self.check_tree_left.column(3, stretch=NO, width=50,anchor=CENTER)

		verscrlbar = ttk.Scrollbar(self.left_frame,
                           orient ="vertical",
                           command = self.check_tree_left.yview)
		verscrlbar.pack( side=RIGHT, fill=Y )
		self.check_tree_left.configure(yscrollcommand=verscrlbar.set)

	def initRightTree(self):
		self.check_tree_right.heading(1, anchor=CENTER, text="#" )
		self.check_tree_right.heading(2, anchor=W, text="File name")
		self.check_tree_right.heading(3, anchor=CENTER, text="Type")
		self.check_tree_right.column(1, stretch=NO, width=25,anchor=CENTER)
		self.check_tree_right.column(2, anchor=W)
		self.check_tree_right.column(3, stretch=NO, width=50,anchor=CENTER)

		verscrlbar = ttk.Scrollbar(self.right_frame,
                           orient ="vertical",
                           command = self.check_tree_right.yview)
		verscrlbar.pack( side=RIGHT, fill=Y )
		self.check_tree_right.configure(yscrollcommand=verscrlbar.set)


	def resetCheckTree(self):
		for child in self.check_tree_left.get_children():
			self.check_tree_left.delete(child)
		for child in self.check_tree_right.get_children():
			self.check_tree_right.delete(child)
		return

	def move_selected(self):
		for item in self.check_tree_left.selection():
			selected_item = self.check_tree_left.item(item)
			if selected_item["values"][2] in ["xlsx", "csv"]:
				self.check_tree_right.insert("", "end", values=selected_item["values"])
				self.check_tree_left.delete(item)

		return

	def move_all(self):
		for item in self.check_tree_left.get_children():
			selected_item = self.check_tree_left.item(item)
			if selected_item["values"][2] in ["xlsx", "csv"]:
				self.check_tree_right.insert("", "end", values=selected_item["values"])
				self.check_tree_left.delete(item)		
		return

	def reverse_one(self):
		for item in self.check_tree_right.selection():
			selected_item = self.check_tree_right.item(item)
			self.check_tree_left.insert("", "end", values=selected_item["values"])
			self.check_tree_right.delete(item)

		return

	def reverse_all(self):
		for item in self.check_tree_right.get_children():
			selected_item = self.check_tree_right.item(item)
			self.check_tree_left.insert("", "end", values=selected_item["values"])
			self.check_tree_right.delete(item)		
		return

	def load_check_folder(self):
		dir_path = self.check_folder_path.get()		
		self.resetCheckTree()
		if path.isdir(dir_path):
			files = listdir(dir_path)
			self.paths = files;

			for index in range(len(files)):
				file = files[index]

				if path.isdir( path.join( dir_path, file ) ):
					self.check_tree_left.insert("", "end", values=( index, file, "dir" ))
				else:
					extension = file.split('.')[-1]
					self.check_tree_left.insert("", "end", values=( index, file, extension ))
		else:
			messagebox(title = "Error", message="Invalid directory path")
		return
	def startChecking(self):
		dir_path = self.check_folder_path.get()		
		files = []
		self.compareData = {}

		for item in self.check_tree_right.get_children():
			selected_item = self.check_tree_right.item(item)
			file = selected_item["values"][1]
			files.append({
				"fullpath": path.join(dir_path, file),
				"name": file,
				"extension": selected_item["values"][2]
			})

		for file in files:
			extension = file["extension"]

			self.file_in_progress.set(file["name"])			
			self.checking_progress.set('CHECKED %.2f%s OF'%(0, "%"))
			if extension == "xlsx":
				self.compareXLSX(file["fullpath"], file["name"])
			if extension == "csv":
				self.compareCSV(file["fullpath"], file["name"])
			self.update()
		openFolder = messagebox.askyesno(title="Info", message="Successfully compared data!, Open result folder ?")
		if openFolder:
			startfile( path.join(dir_path, "CHECK_RESULT") )

	def compareXLSX(self, file, name):
		wb = openpyxl.load_workbook(file)
		sheet = wb.sheetnames[0]
		ws = wb[sheet]
		header = self.check_header.get()

		xl = pd.ExcelFile(file)
		df = xl.parse(sheet)	

		cols = list(df.columns)
		
		if header in cols:
			index = cols.index( header )
			df_rows = df.to_numpy().tolist()
			col_letter = get_column_letter(index + 1)
			max_length = len( df_rows )
			for i in range(len(df_rows)):				
				self.checking_progress.set('CHECKED %.2f%% OF'%(i * 100 / max_length))
				row = df_rows[i]
				key = row[index]
				keys = self.compareData.keys()
				if key in keys:
					ws[f'{col_letter}{i + 2}'].font = Font(color="FF0000")
					
				else:
					self.compareData.setdefault( key, 1 )
				self.update()
		
		check_folder_path = self.check_folder_path.get()
		CHECK_RESULT_PATH = path.join( check_folder_path, "CHECK_RESULT" )
		if not path.isdir( CHECK_RESULT_PATH ):
			mkdir(CHECK_RESULT_PATH)

		export_file_name = path.join( CHECK_RESULT_PATH, name )		
		wb.save(export_file_name)
		return

	def compareCSV(self, file, name):
		header = self.check_header.get()	

		data = pd.read_csv(file)
		columns = data.columns.values	
		raw_csv = open( file )
		df = pd.DataFrame([raw_csv], index = None)
		
		max_length = len(data)
		
		for i in range(max_length):
			row = list(df[i+1])
			row = row[0].replace("\n", "")			
			splittedRow = row.split(',')
			serialized = dict()			
			dataSet = set()
			for col in range(len(columns)):
				serialized.setdefault(columns[col], splittedRow[col])						
				dataSet = (*dataSet, splittedRow[col])
			self.checking_progress.set('CHECKED %.2f%% OF'%(i * 100 / max_length))

			
			key = serialized[header]
			keys = self.compareData.keys()
			if key in keys:				
				csv_data.append( list((*dataSet, "DUPLICATED")) )				
			else:
				self.compareData.setdefault( key, 1 )
				csv_data.append( list((*dataSet, "")) )
			self.update()
		
		check_folder_path = self.check_folder_path.get()
		CHECK_RESULT_PATH = path.join( check_folder_path, "CHECK_RESULT" )
		if not path.isdir( CHECK_RESULT_PATH ):
			mkdir(CHECK_RESULT_PATH)

		export_file_name = path.join( CHECK_RESULT_PATH, name )
		df = pd.DataFrame(csv_data)
		df.columns = [*columns, "CHECK_RESULT"]
		df.to_csv(export_file_name)

__all__ = [ 'Window' ]
